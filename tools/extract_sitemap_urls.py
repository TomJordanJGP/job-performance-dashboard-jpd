"""Pull every <loc> URL from the jobsgopublic.com sitemap tree and write to xlsx.

Walks the two sitemap indexes, fetches each child sitemap, extracts URLs and
last-modified timestamps, and writes a single workbook with a 'sitemap' column
flagging which child each URL came from.
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NS = {'sm': 'http://www.sitemaps.org/schemas/sitemap/0.9'}

INDEXES = [
    'https://www.jobsgopublic.com/sitemap.xml',
    'https://www.jobsgopublic.com/search/sitemap.xml',
]

OUTPUT = Path(__file__).resolve().parent.parent / 'jobsgopublic_sitemap_urls.xlsx'


def fetch(url: str) -> bytes:
    resp = requests.get(url, timeout=30, headers={'User-Agent': 'sitemap-export/1.0'})
    resp.raise_for_status()
    return resp.content


def child_sitemaps(index_url: str) -> list[str]:
    root = ET.fromstring(fetch(index_url))
    return [loc.text.strip() for loc in root.findall('sm:sitemap/sm:loc', NS) if loc.text]


def urls_in_sitemap(sitemap_url: str) -> list[tuple[str, str]]:
    root = ET.fromstring(fetch(sitemap_url))
    out: list[tuple[str, str]] = []
    for url_el in root.findall('sm:url', NS):
        loc = url_el.find('sm:loc', NS)
        lastmod = url_el.find('sm:lastmod', NS)
        if loc is not None and loc.text:
            out.append((loc.text.strip(), (lastmod.text.strip() if lastmod is not None and lastmod.text else '')))
    return out


def main() -> int:
    rows: list[tuple[str, str, str, str]] = []  # (sitemap, url, lastmod, source_index)

    for index_url in INDEXES:
        print(f'Reading index: {index_url}', file=sys.stderr)
        children = child_sitemaps(index_url)
        print(f'  -> {len(children)} child sitemap(s)', file=sys.stderr)
        for child in children:
            print(f'  fetching {child}', file=sys.stderr)
            for url, lastmod in urls_in_sitemap(child):
                rows.append((child, url, lastmod, index_url))

    print(f'Total URLs: {len(rows)}', file=sys.stderr)

    wb = Workbook()
    ws = wb.active
    ws.title = 'URLs'

    headers = ['URL', 'Last Modified', 'Child Sitemap', 'Parent Sitemap Index']
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='305496')
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for child, url, lastmod, index_url in rows:
        ws.append([url, lastmod, child, index_url])

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    widths = {1: 80, 2: 28, 3: 60, 4: 60}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    summary = wb.create_sheet('Summary')
    summary.append(['Generated', datetime.now().isoformat(timespec='seconds')])
    summary.append(['Total URLs', len(rows)])
    summary.append([])
    summary.append(['Child Sitemap', 'URL Count'])
    counts: dict[str, int] = {}
    for child, _, _, _ in rows:
        counts[child] = counts.get(child, 0) + 1
    for child, count in counts.items():
        summary.append([child, count])
    summary.column_dimensions['A'].width = 70
    summary.column_dimensions['B'].width = 18
    for cell in summary[4]:
        cell.font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Wrote {OUTPUT}', file=sys.stderr)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
